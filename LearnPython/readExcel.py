import openpyxl
wb = openpyxl.load_workbook('./Files/example.xlsx')
#print(type(wb))
#print(wb.sheetnames)

#sheet = wb['Data1']
sheet = wb.active
print(sheet)
#print(sheet.title)

'''
for i in range(2, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=' ')
    print()
'''

for rowObject in sheet['A2':'C5']:
    for cellObject in rowObject:
        print(cellObject.value, end=' ')
    print()



